#!/usr/bin/env python3
"""
RachelUGC — Extract UGC Tracker spreadsheet to data.json.

Reads ~/Downloads/Rachel/UGC Tracker.xlsx (or --input PATH) and writes
RachelUGC/data.json, the source data the SPA loads on first launch.

Re-run any time Rachel updates her spreadsheet. The SPA detects the
baseline change on next page load and asks whether to keep in-browser
edits or discard them.

Self-bootstraps a .venv with openpyxl so the only system requirement is
python3 >= 3.10.
"""

import argparse
import json
import os
import re
import subprocess
import sys
from datetime import datetime, timezone
from pathlib import Path

__version__ = "0.1.0"

SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
VENV_DIR = SCRIPT_DIR / ".venv"
VENV_PY = VENV_DIR / "bin" / "python"

DEFAULT_INPUT = Path.home() / "Downloads" / "Rachel" / "UGC Tracker.xlsx"
DEFAULT_OUTPUT = PROJECT_DIR / "data.json"


def _in_venv() -> bool:
    return Path(sys.prefix).resolve() == VENV_DIR.resolve()


def _venv_can_import_openpyxl() -> bool:
    if not VENV_PY.is_file():
        return False
    try:
        r = subprocess.run(
            [str(VENV_PY), "-c", "import openpyxl"],
            check=False, capture_output=True, timeout=10,
        )
        return r.returncode == 0
    except (OSError, subprocess.TimeoutExpired):
        return False


def _bootstrap_venv() -> None:
    if not VENV_PY.is_file():
        print(f"Creating venv at {VENV_DIR} ...", file=sys.stderr)
        subprocess.run([sys.executable, "-m", "venv", str(VENV_DIR)], check=True)
    if not _venv_can_import_openpyxl():
        print("Installing openpyxl into venv ...", file=sys.stderr)
        subprocess.run(
            [str(VENV_PY), "-m", "pip", "install", "--quiet",
             "--disable-pip-version-check", "openpyxl>=3.1.0"],
            check=True,
        )
    os.execv(str(VENV_PY), [str(VENV_PY), str(Path(__file__).resolve()), *sys.argv[1:]])


if not _in_venv():
    _bootstrap_venv()


import openpyxl  # noqa: E402


MONTH_ALIASES = {
    "JAN": "January", "FEB": "February", "MAR": "March", "APR": "April",
    "MAY": "May", "JUN": "June", "JUL": "July", "AUG": "August",
    "SEP": "September", "SEPT": "September", "OCT": "October",
    "NOV": "November", "DEC": "December",
}

SUBTOTAL_RE = re.compile(
    r"^\s*(?P<month>[A-Z]+)\s+TOTAL\s+CONTRACTED\s*:\s*\$?(?P<amount>[\d,]+(?:\.\d+)?)",
    re.IGNORECASE,
)

CURRENT_MONTH_NAME = datetime.now().strftime("%B").upper()  # e.g. "MAY"
CURRENT_MONTH_CODE = CURRENT_MONTH_NAME[:3]


def _coerce_amount(raw):
    """Map the spreadsheet's `$$ Amount` cell to (numeric, isGifted, rawDisplay).

    Cells can be:
      - a number (e.g. 240.0) → (240.0, False, "$240")
      - the literal string "Gifted" → (0.0, True, "Gifted")
      - "Gifted/$250 GC" or similar → (0.0, True, raw string)
      - None / blank → (0.0, False, "")
    """
    if raw is None:
        return 0.0, False, ""
    if isinstance(raw, (int, float)):
        return float(raw), False, f"${int(raw) if raw == int(raw) else raw:.2f}".rstrip("0").rstrip(".") if isinstance(raw, float) else f"${raw}"
    s = str(raw).strip()
    if not s:
        return 0.0, False, ""
    lower = s.lower()
    if "gift" in lower:
        # "Gifted", "Gifted/$250 GC", etc.
        return 0.0, True, s
    # Try to coerce a stray numeric-string
    try:
        val = float(s.replace("$", "").replace(",", ""))
        return val, False, f"${val:g}"
    except ValueError:
        return 0.0, False, s


def parse_workbook(path: Path) -> dict:
    """Two-pass parse: collect deals + subtotal markers in source order, then
    back-assign each unlabeled "deal" row to the next subtotal that follows.

    The spreadsheet's structure: a contiguous block of deal rows, followed by
    a subtotal row, followed by the next month's block, etc. The most recent
    (open) month often has a bare-amount-only subtotal row at the very bottom
    instead of a labeled "MONTH TOTAL CONTRACTED" — we map that to the current
    calendar month.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Spreadsheet at {path} is empty.")

    header = [(c or "").strip() if isinstance(c, str) else c for c in rows[0]]
    expected = ["Platform", "Brand Name", "Deliverables", "$$ Amount",
                "Status", "Posting to Socials?", "Paid", "Category"]
    if [str(c).strip() for c in header[:8]] != expected:
        print(f"WARNING: spreadsheet headers don't match expected layout:\n"
              f"  expected: {expected}\n"
              f"  found:    {header[:8]}",
              file=sys.stderr)

    # Pass 1: classify each row as deal | labeled-subtotal | bare-subtotal | skip
    items: list[dict] = []
    seq = 0
    for i, row in enumerate(rows[1:], start=2):
        platform, brand, deliverables, amount, status, posting, paid, category = (
            (list(row) + [None] * 8)[:8]
        )

        all_blank = all(
            c is None or (isinstance(c, str) and not c.strip())
            for c in row
        )
        if all_blank:
            continue

        # Labeled subtotal: "MARCH TOTAL CONTRACTED: $243"
        if not platform and isinstance(brand, str):
            m = SUBTOTAL_RE.match(brand)
            if m:
                month_code = m.group("month").upper()[:3]  # normalise to 3-letter code
                items.append({
                    "kind": "subtotal",
                    "rowIndex": i,
                    "code": month_code,
                    "name": MONTH_ALIASES.get(month_code, m.group("month").title()),
                    "contracted": float(m.group("amount").replace(",", "")),
                })
                continue

        # Bare-amount subtotal: only $$ Amount populated. Assigned to the
        # current calendar month — this is Rachel's running tally for the
        # month-in-progress that she hasn't given a label to yet.
        if (not platform and not brand and not deliverables and not status
                and isinstance(amount, (int, float))):
            items.append({
                "kind": "subtotal",
                "rowIndex": i,
                "code": CURRENT_MONTH_CODE,
                "name": CURRENT_MONTH_NAME.title(),
                "contracted": float(amount),
                "bare": True,
            })
            continue

        # Deal row.
        if not platform and not brand:
            continue
        seq += 1
        amount_numeric, is_gifted, amount_display = _coerce_amount(amount)
        items.append({
            "kind": "deal",
            "deal": {
                "id": f"row-{seq:03d}",
                "rowIndex": i,
                "month": None,  # filled in below
                "platform": (str(platform).strip() if platform else ""),
                "brand": (str(brand).strip() if brand else ""),
                "deliverables": (str(deliverables).strip() if deliverables else ""),
                "amountRaw": str(amount) if amount is not None else "",
                "amountDisplay": amount_display,
                "amountNumeric": amount_numeric,
                "isGifted": is_gifted,
                "status": (str(status).strip() if status else ""),
                "postingToSocials": (str(posting).strip() if posting else ""),
                "paid": (str(paid).strip() if paid else ""),
                "category": (str(category).strip() if category else ""),
            },
        })

    # Pass 2: each subtotal "owns" deals preceding it (back to the previous
    # subtotal). Walk forward through items, batching deals until we hit a
    # subtotal, then flush the batch with that subtotal's month code.
    months: list[dict] = []
    deals: list[dict] = []
    pending: list[dict] = []
    for it in items:
        if it["kind"] == "deal":
            pending.append(it["deal"])
        else:
            # Subtotal — flush pending deals into this month
            for d in pending:
                d["month"] = it["code"]
            month = {
                "code": it["code"],
                "name": it["name"],
                "contracted": it["contracted"],
                "rowIndex": it["rowIndex"],
                "deals": [d["id"] for d in pending],
                "bare": it.get("bare", False),
            }
            months.append(month)
            deals.extend(pending)
            pending = []

    # Any leftover pending deals never hit a subtotal — assign to current month
    # but don't synthesise a month bucket if there's already one for it.
    if pending:
        for d in pending:
            d["month"] = CURRENT_MONTH_CODE
        # If current-month bucket already exists, just append; else create.
        existing = next((m for m in months if m["code"] == CURRENT_MONTH_CODE), None)
        if existing:
            existing["deals"].extend(d["id"] for d in pending)
        else:
            months.append({
                "code": CURRENT_MONTH_CODE,
                "name": CURRENT_MONTH_NAME.title(),
                "contracted": sum(d["amountNumeric"] for d in pending),
                "deals": [d["id"] for d in pending],
                "bare": True,
            })
        deals.extend(pending)

    computed_total = sum(d["amountNumeric"] for d in deals)

    return {
        "extractedAt": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "source": str(path),
        "computedTotal": computed_total,
        "months": months,
        "deals": deals,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--input", "-i", type=Path, default=DEFAULT_INPUT,
                    help=f"path to UGC Tracker.xlsx (default: {DEFAULT_INPUT})")
    ap.add_argument("--output", "-o", type=Path, default=DEFAULT_OUTPUT,
                    help=f"output JSON path (default: {DEFAULT_OUTPUT})")
    ap.add_argument("--print-summary", action="store_true",
                    help="print a one-line summary to stderr after writing")
    args = ap.parse_args()

    if not args.input.is_file():
        print(f"ERROR: input spreadsheet not found at {args.input}", file=sys.stderr)
        return 2

    payload = parse_workbook(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n")

    # Also emit data.js next to data.json. The SPA loads this via
    # <script src="data.js"> because file:// blocks fetch() in most browsers,
    # so we can't fetch("data.json") when Rachel just double-clicks index.html.
    js_path = args.output.with_suffix(".js")
    js_path.write_text(
        "// Auto-generated by scripts/extract.py — do not edit by hand.\n"
        "// Re-run `python3 scripts/extract.py` to refresh.\n"
        "window.BOOTSTRAP_DATA = "
        + json.dumps(payload, indent=2, ensure_ascii=False)
        + ";\n"
    )

    if args.print_summary or sys.stdout.isatty():
        n_deals = len(payload["deals"])
        n_months = len(payload["months"])
        ct = payload["computedTotal"]
        print(f"Wrote {args.output}", file=sys.stderr)
        print(f"  {n_deals} deals across {n_months} months", file=sys.stderr)
        for m in payload["months"]:
            marker = " (open)" if m.get("bare") else ""
            print(f"    {m['code']:>4} {m['name']:<10} "
                  f"contracted=${m['contracted']:.2f} "
                  f"deals={len(m['deals'])}{marker}",
                  file=sys.stderr)
        print(f"  YTD computed total: ${ct:.2f}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
